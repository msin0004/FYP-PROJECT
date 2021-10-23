# Program: SUMOPaint
# Goal: Permits the user to connect to the Monash SUMOPaint simulation and subsequently paint the vehicle of interest any desired colour.
# Author: Wynita Griggs
# Editor: Meher Singh
# Date: 20 July, 2021
# Tested and works with SUMO 1.8.0.

import os, sys
import traci
import multiprocessing
from multiprocessing import Process, Value, Manager
from ctypes import c_wchar_p
import socket
import time
import threading
#for excel saving
import xlwt
from openpyxl import load_workbook
sending_data = "sample"


# server program:  creates the server, handles incoming calls and subsequent user requests
def server(data,temp,central_authority_broadcast, bus_send, eta_send, pass_send, bus_no, stop, list, bus_info, received, solution):
	# size of buffer and backlog
	buffer = 2048 # value should be a relatively small power of 2, e.g. 4096
	backlog = 1 # tells the operating system to keep a backlog of 1 connection; this means that you can have at most 1 client waiting while the server is handling the current client; the operating system will typically allow a maximum of 5 waiting connections; to cope with this, busy servers need to generate a new thread to handle each incoming connection so that it can quickly serve the queue of waiting clients

	# create a socket
	server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM) # AF_INET = IPv4 socket family; SOCK_STREAM = TCP socket type

	# bind the socket to an address and port
	host = '127.0.0.1' # localhost
	port = 8080 # reserve a port for the service (i.e. a large number less than 2^16); the call will fail if some other application is already using this port number on the same machine
	server_socket.bind((host, port)) # binds the socket to the hostname and port number

	# listen for incoming connections
	server_socket.listen(backlog)
	realVehicleIndex = 0


	def clientHandle(client_socket, address, realVehicleIndex):
		testing = 1 #for testing purposes
		sent_flag = 0 #for when the user sends the final values
		while True: # infinite loop 2
			incoming = client_socket.recv(buffer).decode('UTF-8') # receive client data into buffer
			print("Receiving data from client.")
			#print(incoming)

			#if client ends session
			if (incoming == 'quit'):
				print("Ending session with client.")
				client_socket.close() # close the connection with the client
				break # breaks out of infinite loop 2

			#if client finishes the solution
			if (received['sent_flag'] == 1 and sent_flag == 0):
				print("Client sent over user decision making choices.")
				solution.value = incoming
				sent_flag = 1
				received['end'] = 1

			#sending bus information code
			if (incoming == 'bus_no'):
				temp.value = ''
				received['operation'] = ''
				print("Client requested bus information for stop " + str(received['stop']) + ".")
				message = str(bus_no.value) + '\n'
				client_socket.send(message.encode())
				if bus_no.value == 1:
					#print("inside sending for 1 bus")
					#send bus 1 eta
					message = str(bus_info['bus1_eta']) + '\n'
					#message = '1\n'
					client_socket.send(message.encode())
					#send bus 1 passengers
					message = str(bus_info['bus1_pass']) + '\n'
					#message = '1\n'
					client_socket.send(message.encode())
					#send bus 1 seats
					message = str(bus_info['bus1_seat']) + '\n'
					#message = '1\n'
					client_socket.send(message.encode())
					#print("ending send for 1 bus")
				elif bus_no.value == 2:
					#print("inside sending for 2 bus")
					#send bus 1 eta
					message = str(bus_info['bus1_eta']) + '\n'
					#message = '2\n'
					client_socket.send(message.encode())
					#send bus 1 passengers
					message = str(bus_info['bus1_pass']) + '\n'
					client_socket.send(message.encode())
					#send bus 1 seats
					message = str(bus_info['bus1_seat']) + '\n'
					client_socket.send(message.encode())
					#send bus 2 eta
					message = str(bus_info['bus2_eta']) + '\n'
					client_socket.send(message.encode())
					#send bus 2 passengers
					message = str(bus_info['bus2_pass']) + '\n'
					client_socket.send(message.encode())
					#send bus 2 seats
					message = str(bus_info['bus2_seat']) + '\n'
					client_socket.send(message.encode())
					#print("ending send for 2 bus")
				elif bus_no.value == 3:
					#print("inside sending for 3 bus")
					#send bus 1 eta
					message = str(bus_info['bus1_eta']) + '\n'
					client_socket.send(message.encode())
					#send bus 1 passengers
					message = str(bus_info['bus1_pass']) + '\n'
					client_socket.send(message.encode())
					#send bus 1 seats
					message = str(bus_info['bus1_seat']) + '\n'
					client_socket.send(message.encode())
					#send bus 2 eta
					message = str(bus_info['bus2_eta']) + '\n'
					client_socket.send(message.encode())
					#send bus 2 passengers
					message = str(bus_info['bus2_pass']) + '\n'
					client_socket.send(message.encode())
					#send bus 2 seats
					message = str(bus_info['bus2_seat']) + '\n'
					client_socket.send(message.encode())
					#send bus 3 eta
					message = str(bus_info['bus3_eta']) + '\n'
					client_socket.send(message.encode())
					#send bus 3 passengers
					message = str(bus_info['bus3_pass']) + '\n'
					client_socket.send(message.encode())
					#send bus 3 seats
					message = str(bus_info['bus3_seat']) + '\n'
					client_socket.send(message.encode())
					#print("ending send for 3 bus")


			if (incoming == 'Stop 7'):
				#stop = 'stop7'
				#print("inside stop 7")
				stop.value = 7
				temp.value = 'bus_no'
				received['stop'] = 7
				received['operation'] = 'bus_no'
			if (incoming == 'Stop 1'):
				#stop = 'stop1'
				#print("inside stop 1")
				stop.value = 1
				temp.value = 'bus_no'
				received['stop'] = 1
				received['operation'] = 'bus_no'
			if (incoming == 'Stop 2'):
				#stop = 'stop2'
				#print("inside stop 2")
				stop.value = 2
				temp.value = 'bus_no'
				received['stop'] = 2
				received['operation'] = 'bus_no'
			if (incoming == 'Stop 3'):
				#stop = 'stop3'
				#print("inside stop 3")
				stop.value = 3
				temp.value = 'bus_no'
				received['stop'] = 3
				received['operation'] = 'bus_no'
			if (incoming == 'Stop 4'):
				#stop = 'stop4'
				#print("inside stop 4")
				stop.value = 4
				temp.value = 'bus_no'
				received['stop'] = 4
				received['operation'] = 'bus_no'
			if (incoming == 'Stop 5'):
				#stop = 'stop5'
				#print("inside stop 5")
				stop.value = 5
				temp.value = 'bus_no'
				received['stop'] = 5
				received['operation'] = 'bus_no'
			if (incoming == 'Stop 6'):
				#stop = 'stop6'
				#print("inside stop 6")
				stop.value = 6
				temp.value = 'bus_no'
				received['stop'] = 6
				received['operation'] = 'bus_no'
			if (incoming == 'Stop 8'):
				#stop = 'stop8'
				#print("inside stop 8")
				stop.value = 8
				temp.value = 'bus_no'
				received['stop'] = 8
				received['operation'] = 'bus_no'

			#FINAL SOLUTION FROM CLIENT
			if (incoming == 'Bus 1'):
				received['sent_flag'] = 1
				received['bus'] = incoming
			if (incoming == 'Bus 2'):
				received['sent_flag'] = 1
				received['bus'] = incoming
			if (incoming == 'Bus 3'):
				received['sent_flag'] = 1
				received['bus'] = incoming



	while True: # infinite loop 1
		client_socket, address = server_socket.accept() # passively accept TCP client connections; the call returns a pair of arguments: client is a new Socket object used to communicate with the client and address is the address of the client

		# record client connection time (as seen from the server)
		start_time = time.strftime('%d %b %Y at %H:%M:%S')
		init_time = str(start_time) # convert connection time to a string
		print('Made a connection with', address, 'on', init_time + '.')

		#threading.newthread(clientHandle, (client_socket, address, realVehicleIndex))
		print("Starting the server thread")
		threading.Thread(target=clientHandle, args=(client_socket, address, realVehicleIndex)).start()


def eta_code(bus_lane, stop, next_stop):
	#eta initialise
	eta = -1

	#lanes per stop
	stop1_2 = [ '787448090#0', '787448090#1', '807426789', '-638398362#2', '-638398362#1', '-638398362#0', '814105115#0', '814105115#1', '814105114', '198133694#0', '198133694#1', '198133694#2', '198133694#3', '198133694#4', '198134599#0', '198134599#1', '198134598#0', '198134598#1', '198134598#2', '198134598#3', '782399755', '144987843#0', '144987843#1', '180708618', '211260230', '180708619', '782395452#0', '782395452#1', '713493692#0', '713493692#1', '791668582', '4419370#0', '4419370#1', '4419370#2', '4419370#3', '4419370#4', '15475366#0', '15475366#1', '48718672', '167571335', '48718680#0', '48718680#1', '48718680#2', '167571359#0', '167571359#1', '198692245#0', '198692245#1', '88463238#2', '88463238#3', '88463238#4', '88463238#5', '88463238#6', '45022884#0', '45022884#1']
	stop2_3 = ['229376392#1', '11588123#0', '11588123#1', '11588123#2', '11588123#3', '11588123#4', '-229376392#1', '-229376392#0', '229376385#0', '229376385#1', '88463238#1', '88463238#2', '88463238#3', '229376387#0', '229376387#1', '-66793003#1', '-66793003#0', '-45023071', '-23714110#1', '-23714110#0', '45023005#2', '45023005#3', '4868063#0', '4868063#1', '4868063#2', '178932596#0', '178932596#1', '178932594#0']
	stop3_4 = ['178931858#0', '178931858#1', '-178933894#1', '-178933894#0', '178931861#0', '178931861#1', '-178933462#1', '-178933462#0', '178931860#0', '178931860#1', '-178933463#1', '-178933463#0', '178931857#0', '178931857#1', '178933466#0', '178933466#1']
	stop4_5 = ['165171774#0', '165171774#1' '165171774#2', '165171774#3', '165171774#4', '165171774#5', '165171774#6', '165171774#7', '165171774#8', '165171774#9', '165171774#10', '165171774#11', '165171774#12', '165171774#13', '48333890', '679193453', '679193452', '679193449']
	stop5_6 = ['679469626#1', '679469626#2', '679193451', '679193452', '679469629', '4612649', '791849036', '202904759', '763416697#0', '763416697#1', '763416697#2']
	stop6_7 = ['198691699#1', '198691699#2', '198691699#3', '792597607#0', '792597607#1', '794258388', '792597606', '573676801', '794258389#0', '794258389#1', '794258389#2']
	stop7_8 = ['794269737', '792153597', '767369896#0', '767369896#1', '767369896#2', '237508424', '139327105#0', '139327105#1', '167571364', '44348023#0', '44348023#1', '139327110', '10539028#0', '10539028#1', '493547206#0', '493547206#1', '807418425#0']
	#lane for every bus stop
	stop1 = '-787851477'
	stop2 = '229376392#0'
	stop3 = '178932594#1'
	stop4 = '165171774#0'
	stop5 = '679469626#0'
	stop6 = '198691699#0'
	stop7 = '198691696'
	stop8 = '807418425#1'

	#stores calculated time between stops in minutes
	time1_2 = 3 	#from 30 seconds to 200 seconds
	time2_3 = 2 	#from 220 seconds to 328 seconds
	time3_4 = 2		#from 348 secodns to 427 seconds
	time4_5 = 3		#from 447 seconds to  595
	time5_6 = 2		#from 615 seconds to 717
	time6_7 = 1		#from 737 seconds to 803
	time7_8 = 2		#from 823 seconds to 914

	#next stop from bus
	time_to_next_stop = 0
	#if found bus between stop (flag)
	find = 0

	#run for each lane between stops
	#stop1-2
	count = 0
	for ind in stop1_2:
		count = count + 1
		#if a lane is equal to the stop
		if bus_lane == ind:
			#get number of lanes reamining till next stop
			pos_rem = len(stop1_2) - count
			#get time to next stop
			time_to_next_stop = pos_rem*(time1_2/len(stop1_2))
			bus_lane = stop2
			find = 1
			break
	#stop2-3
	count = 0
	if find == 0:
		for ind in stop2_3:
			count = count + 1
			#if a lane is equal to the stop
			if bus_lane == ind:
				#get number of lanes reamining till next stop
				pos_rem = len(stop2_3) - count
				#get time to next stop
				time_to_next_stop = pos_rem*(time2_3/len(stop2_3))
				bus_lane = stop3
				find = 1
				break
	#stop3-4
	count = 0
	if find == 0:
		for ind in stop3_4:
			count = count + 1
			#if a lane is equal to the stop
			if bus_lane == ind:
				#get number of lanes reamining till next stop
				pos_rem = len(stop3_4) - count
				#get time to next stop
				time_to_next_stop = pos_rem*(time3_4/len(stop3_4))
				bus_lane = stop4
				find = 1
				break

	#stop4-5
	count = 0
	if find == 0:
		for ind in stop4_5:
			count = count + 1
			#if a lane is equal to the stop
			if bus_lane == ind:
				#get number of lanes reamining till next stop
				pos_rem = len(stop4_5) - count
				#get time to next stop
				time_to_next_stop = pos_rem*(time4_5/len(stop4_5))
				bus_lane = stop5
				find = 1
				break
	#stop5-6
	count = 0
	if find == 0:
		for ind in stop5_6:
			count = count + 1
			#if a lane is equal to the stop
			if bus_lane == ind:
				#get number of lanes reamining till next stop
				pos_rem = len(stop5_6) - count
				#get time to next stop
				time_to_next_stop = pos_rem*(time5_6/len(stop5_6))
				bus_lane = stop6
				find = 1
				break
	#stop6-7
	count = 0
	if find == 0:
		for ind in stop6_7:
			count = count + 1
			#if a lane is equal to the stop
			if bus_lane == ind:
				#get number of lanes reamining till next stop
				pos_rem = len(stop6_7) - count
				#get time to next stop
				time_to_next_stop = pos_rem*(time6_7/len(stop6_7))
				bus_lane = stop7
				find = 1
				break
	#stop7-8
	count = 0
	if find == 0:
		for ind in stop7_8:
			count = count + 1
			#if a lane is equal to the stop
			if bus_lane == ind:
				#get number of lanes reamining till next stop
				pos_rem = len(stop7_8) - count
				#get time to next stop
				time_to_next_stop = pos_rem*(time7_8/len(stop7_8))
				bus_lane = stop8
				find = 1
				break

	if find == 0:
		if bus_lane != stop1 and bus_lane != stop2 and bus_lane != stop3 and bus_lane != stop4 and bus_lane != stop5 and bus_lane != stop6 and bus_lane != stop7 and bus_lane != stop8:
			if next_stop == 2:
				time_to_next_stop = time1_2/2
				bus_lane = stop2
			if next_stop == 3:
				time_to_next_stop = time2_3/2
				bus_lane = stop3
			if next_stop == 4:
				time_to_next_stop = time3_4/2
				bus_lane = stop4
			if next_stop == 5:
				time_to_next_stop = time4_5/2
				bus_lane = stop5
			if next_stop == 6:
				time_to_next_stop = time5_6/2
				bus_lane = stop6
			if next_stop == 7:
				time_to_next_stop = time6_7/2
				bus_lane = stop7
			if next_stop == 8:
				time_to_next_stop = time7_8/2
				bus_lane = stop8

	#print(time_to_next_stop)
	#check if bus is at a current stop
	if bus_lane == stop1:
		#print("STOP1")
		#if check every stop
		if stop == 1:
			eta = 0
		if stop == 2:
			eta = time1_2 + time_to_next_stop
		if stop == 3:
			eta = time1_2 + time2_3 + time_to_next_stop
		if stop == 4:
			eta = time1_2 + time2_3 + time3_4 + time_to_next_stop
		if stop == 5:
			eta = time1_2 + time2_3 + time3_4 + time4_5 + time_to_next_stop
		if stop == 6:
			eta = time1_2 + time2_3 + time3_4 + time4_5 + time5_6 + time_to_next_stop
		if stop == 7:
			eta = time1_2 + time2_3 + time3_4 + time4_5 + time5_6 + time6_7 + time_to_next_stop
		if stop == 8:
			eta = time1_2 + time2_3 + time3_4 + time4_5 + time5_6 + time6_7 + time7_8 + time_to_next_stop

	elif bus_lane == stop2:
		#print("STOP2")
		#if check every stop
		if stop == 2:
			eta = 0 + time_to_next_stop
		if stop == 3:
			eta = time2_3 + time_to_next_stop
		if stop == 4:
			eta = time2_3 + time3_4 + time_to_next_stop
		if stop == 5:
			eta = time2_3 + time3_4 + time4_5 + time_to_next_stop
		if stop == 6:
			eta = time2_3 + time3_4 + time4_5 + time5_6 + time_to_next_stop
		if stop == 7:
			eta = time2_3 + time3_4 + time4_5 + time5_6 + time6_7 + time_to_next_stop
		if stop == 8:
			eta = time2_3 + time3_4 + time4_5 + time5_6 + time6_7 + time7_8 + time_to_next_stop
	elif bus_lane == stop3:
		#print("STOP3")
		time_to_next_stop = 0
		#if check every stop
		if stop == 3:
			eta = 0 + time_to_next_stop
		if stop == 4:
			eta = time3_4 + time_to_next_stop
		if stop == 5:
			eta = time3_4 + time4_5 + time_to_next_stop
		if stop == 6:
			eta = time3_4 + time4_5 + time5_6 + time_to_next_stop
		if stop == 7:
			eta = time3_4 + time4_5 + time5_6 + time6_7 + time_to_next_stop
		if stop == 8:
			eta = time3_4 + time4_5 + time5_6 + time6_7 + time7_8 + time_to_next_stop
	elif bus_lane == stop4:
		#print("STOP4")
		#if check every stop
		if stop == 4:
			eta = 0 + time_to_next_stop
		if stop == 5:
			eta = time4_5 + time_to_next_stop
		if stop == 6:
			eta = time4_5 + time5_6 + time_to_next_stop
		if stop == 7:
			eta = time4_5 + time5_6 + time6_7 + time_to_next_stop
		if stop == 8:
			eta = time4_5 + time5_6 + time6_7 + time7_8 + time_to_next_stop
	elif bus_lane == stop5:
		#print("STOP5")
		#if check every stop
		if stop == 5:
			eta = 0 + time_to_next_stop
		if stop == 6:
			eta = time5_6 + time_to_next_stop
		if stop == 7:
			eta = time5_6 + time6_7 + time_to_next_stop
		if stop == 8:
			eta = time5_6 + time6_7 + time7_8 + time_to_next_stop
	elif bus_lane == stop6:
		#print("STOP6")
		#if check every stop
		if stop == 6:
			eta = 0 + time_to_next_stop
		if stop == 7:
			eta = time6_7 + time_to_next_stop
		if stop == 8:
			eta = time6_7 + time7_8 + time_to_next_stop
	elif bus_lane == stop7:
		#print("STOP7")
		#if check every stop
		if stop == 7:
			eta = 0 + time_to_next_stop
		if stop == 8:
			eta = time7_8 + time_to_next_stop
	elif bus_lane == stop8:
		#print("STOP8")
		#if check every stop
		eta = -1
		if stop == 8:
			eta = 0 + time_to_next_stop

	return eta

# main program
if __name__ == '__main__':

	# constants
	endSim = 3600000 # the simulation will be permitted to run for a total of endSim milliseconds; 1800000 = 30 minutes, 3600000 = 1hr
	timeout = 1 #controls the speed of the simulation with 1 = 1 second, 2 = 0.5 seconds, 0.5 = 2 seconds, 0.1

	# initialisations
	step = 0 # time step
	d = Value('d', 0.0) # 'd' is a string containing a type code as used by the array module (where 'd' is a floating point number implemented in double in C) and 0.0 is an initial value for 'd'
	#d = 'temp'

	manager = Manager()
	central_authority_broadcast = manager.list()
	temp = manager.Value(c_wchar_p, "")
	bus_send = manager.Value(c_wchar_p, "")
	eta_send = manager.Value(c_wchar_p, "")
	pass_send = manager.Value(c_wchar_p, "")
	bus_no = manager.Value('d',0)
	stop = manager.Value('d',0)
	list = manager.list()
	bus_info = manager.dict()  						#stores data obtained from sumo
	received = manager.dict()						#stores data received from client
	solution = manager.Value(c_wchar_p, "") 		#stores the final solution
	received['sent_flag'] = 0						#flag that signals client sending the decision
	received['end'] = 0
	received['operation'] = ''
	bus_info['bus1_eta'] = ''
	bus_info['bus1_pass'] = ''
	bus_info['bus1_seat'] = ''
	bus_info['bus2_eta'] = ''
	bus_info['bus2_pass'] = ''
	bus_info['bus2_seat'] = ''
	bus_info['bus3_eta'] = ''
	bus_info['bus3_pass'] = ''
	bus_info['bus3_seat'] = ''
	bus_info['no_stops'] = ''

	thread = Process(target=server, args=(d, temp, central_authority_broadcast, bus_send, eta_send, pass_send, bus_no, stop, list, bus_info, received, solution)) # represents a task (i.e. the server program) running in a subprocess

	#manager = Manager()
	#central_authority_broadcast = manager.list()
	#thread = Process(target=server, args=(d,central_authority_broadcast)) # represents a task (i.e. the server program) running in a subprocess


	print()
	print('===========================')
	print('Beginning the main program.')
	print('===========================')
	print()
	print("Connecting to SUMO via TraCI.")
	print()

	# import TraCI (to use the library, the <SUMO_HOME>/tools directory must be on the python load path)
	if 'SUMO_HOME' in os.environ:
		tools = os.path.join(os.environ['SUMO_HOME'], 'tools')
		sys.path.append(tools)
	else:
		sys.exit("Please declare environment variable 'SUMO_HOME'.")

	# compose the command line to start SUMO-GUI
	sumoBinary = "/Program Files (x86)/Eclipse/Sumo/bin/sumo-gui"
	sumoCmd = [sumoBinary, "-S", "-c", "bus_route.sumocfg"]
	# start the simulation and connect to it with the python script
	traci.start(sumoCmd)


	print("Launching the server.")
	thread.start()
	print("The server has been launched.")

	# OPEN THE EXCEL Workbook
	book = load_workbook('Bus_Decisions_Data_Gathered.xlsx')
	#book = xlwt.Workbook()
	save_time = str(time.strftime('%d %b %Y at %H.%M.%S'))
	#create new sheet with current date/time
	sheet = book.create_sheet(save_time)


	##	LOCATION OF ALL VARIABLES CREATED
	##
	flag_stops = 1  			#flag to check if code has run before in the get next stop section
	occupied = 1				#current people on the bus
	capacity = 1				#bus Capacity
	#bus_id = 'bus_flow.0'		#temporary variable to store bus id for data collection (need to change to allow dynamic readings)
	t = 0
	end_excel = 0
	prev_eta1 = 0
	prev_eta2 = 0
	prev_eta3 = 0

	while step < endSim:
		thread.join(timeout) # implicitly controls the speed of the simulation; blocks the main program either until the server program terminates (if no timeout is defined) or until the timeout occurs

		#lets the simulation run for 3 seconds before any command can run
		#to let the simulation initialise first
		if step > 3000:
			##GET NUMBER OF BUSSES FOR THE GIVEN STOP
			#VARIABLES BEING used
			#temp.value, stop.value, bus_no.value
			#this gets the road of the current bus
			#this gets the number of bus stops
			#test3 = traci.busstop.getIDCount()

			bus_id_all = traci.vehicle.getIDList()
			stop_num = 0
			count_bus_no = 0
			eta = -1
			#check what the device is asking for
			if received['operation'] == 'bus_no':
				#loop that runs for each simulated bus
				for num in bus_id_all:
					#getting bus data (next stop data)
					bus_data = traci.vehicle.getStops(num, 1)
					next_stop = bus_data[0].stoppingPlaceID
					#converting next stop to number for comparison
					if next_stop == 'r1s1':
						stop_num = 1
					if next_stop == 'r1s2':
						stop_num = 2
					if next_stop == 'r1s3':
						stop_num = 3
					if next_stop == 'r1s4':
						stop_num = 4
					if next_stop == 'r1s5':
						stop_num = 5
					if next_stop == 'r1s6':
						stop_num = 6
					if next_stop == 'r1s7':
						stop_num = 7
					if next_stop == 'r1s8':
						stop_num = 8
					#comparing simulated bus stop to the device requested stop
					if stop_num <= stop.value and count_bus_no < 3:
						#incremenet bus count
						count_bus_no = count_bus_no + 1
						#put code to get data for each bus here
						#USE the following variables
						#bus limit = 64 passengers
						#bus seats = 38 seats

						#forever loop gets stuck
						#while eta == -1:
						road_id = traci.vehicle.getRoadID(num)
						eta = eta_code(road_id, stop.value, stop_num)

						if count_bus_no == 1:
							if eta == -1:
								eta = prev_eta1
							prev_eta1 = eta
							bus_info['bus1_eta'] = str(int(eta))
							bus_info['bus1_pass'] = str(traci.vehicle.getPersonNumber(num)) + '/64'
							bus_info['bus1_seat'] = str(traci.vehicle.getPersonNumber(num)) + '/38'
						if count_bus_no == 2:
							if eta == -1:
								eta = prev_eta2
							prev_eta2 = eta
							bus_info['bus2_eta'] = str(int(eta))
							bus_info['bus2_pass'] = str(traci.vehicle.getPersonNumber(num)) + '/64'
							bus_info['bus2_seat'] = str(traci.vehicle.getPersonNumber(num)) + '/38'
						if count_bus_no == 3:
							if eta == -1:
								eta = prev_eta3
							prev_eta3 = eta
							bus_info['bus3_eta'] = str(int(eta))
							bus_info['bus3_pass'] = str(traci.vehicle.getPersonNumber(num)) + '/64'
							bus_info['bus3_seat'] = str(traci.vehicle.getPersonNumber(num)) + '/38'




				#end of for loop
				#send the count to the server variable
				bus_no.value = count_bus_no
				#print(bus_no.value)
				bus_info['no_of_bus'] = count_bus_no

			#end of if statement for bus number


			if(received['end'] == 1 and end_excel == 0):
				#create excel sheet for storage
				#print("Received Solution from user\n")
				#print(str(solution.value))
				#print("Bus picked by user\n")
				#print(str(received['bus']))

				data = 'Data gathered from session with app'
				sheet['A1'] = data

				if bus_no.value >= 1:
					sheet['A3'] = 'Bus 1'
					sheet['A4'] = 'ETA'
					sheet['B4'] = bus_info['bus1_eta']
					sheet['A5'] = 'Passengers Onboard'
					sheet['B5'] = bus_info['bus1_pass']
					sheet['A6'] = 'Seats Taken'
					sheet['B6'] = bus_info['bus1_seat']
				if bus_no.value >= 2:
					sheet['A8'] = 'Bus 2'
					sheet['A9'] = 'ETA'
					sheet['B9'] = bus_info['bus2_eta']
					sheet['A10'] = 'Passengers Onboard'
					sheet['B10'] = bus_info['bus2_pass']
					sheet['A11'] = 'Seats Taken'
					sheet['B11'] = bus_info['bus2_seat']
				if bus_no.value >= 3:
					sheet['A13'] = 'Bus 3'
					sheet['A14'] = 'ETA'
					sheet['B14'] = bus_info['bus3_eta']
					sheet['A15'] = 'Passengers Onboard'
					sheet['B15'] = bus_info['bus3_pass']
					sheet['A16'] = 'Seats Taken'
					sheet['B16'] = bus_info['bus3_seat']

				sentence = 'The following response was obtained from the app'
				sheet['A18'] = sentence
				bus_stop = 'Bus stop that the user is at'
				sheet['A19'] = bus_stop
				sheet['B19'] = 'Stop ' + str(received['stop'])
				bus_selected = 'Bus selected by the user'
				sheet['A20'] = bus_selected
				sheet['B20'] = received['bus']
				decision = 'User reasoning behing the bus decision'
				sheet['A21'] = decision
				sheet['B21'] = solution.value		#Print out the user output

				#print out data for stop selected, bus selected as well

				book.save('Bus_Decisions_Data_Gathered.xlsx')
				end_excel = 1		#Change flag to get out of loop
				print("Information obtained with this client session has been stored in the excel database.")


		# go to the next time step
		step += 1000 # in milliseconds
		traci.simulationStep()

	print("Shutting the server down.")
	thread.terminate()
	print("Closing the main program. Goodbye.")
	traci.close() # close the connection to SUMO
